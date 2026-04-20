import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import io

# --- 1. Configuración de Constantes y Nombres ---
HOJA_ACTUAL = 'Costos ACTUAL'
HOJA_ANTERIOR = 'Costos ANTERIOR'
HOJA_PROCESADA = 'Costos_procesado'
HOJA_CONSOLIDADO = 'Consolidado_Impactos'

COLUMNA_RESULTADO = 'Result'
CLAVE_MERGE = 'Material'
NOMBRES_COSTOS_INTERNOS = ['Marteri', 'Materia_Costo', 'Alistam', 'Mano de', 'Maquila', 'Energ', 'Maqui', 'Cif']

# Columnas que deben ser ENTEROS (Redondeo a 0 decimales)
COLUMNAS_ENTEROS = [
    'Marteri', 'Materia_Costo', 'Alistam', 'Mano de', 
    'Maquila', 'Energ', 'Maqui', 'Cif'
]

# Mapeo de nombres internos a nombres de salida (para encabezados)
output_cost_names = {
    'Marteri': 'Marteri', 
    'Materia_Costo': 'Material d', 
    'Alistam': 'Alistam', 
    'Mano de': 'Mano de', 
    'Maquila': 'Maquila', 
    'Energ': 'Energ', 
    'Maqui': 'Maqui', 
    'Cif': 'Cif'
}

# Definición del tipo de borde para los bloques de cálculo
side_medium = Side(border_style='medium', color="000000")
border_left = Border(left=side_medium)
border_right = Border(right=side_medium)

# --- FUNCIÓN 1: Aplicar Formato a la Hoja Procesada (AJUSTADA PARA USAR WORKBOOK) ---
def apply_excel_formatting(wb, sheet_name):
    try:
        # --- Configuración de Estilos ---
        fill_actual_orange = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        fill_variacion_blue = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        fill_impacto_green = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
        font_black_bold = Font(color="000000", bold=True)
        currency_format = '#,##0'

        if sheet_name not in wb.sheetnames:
            st.warning(f"Advertencia: La hoja '{sheet_name}' no se encontró para aplicar formato.")
            return

        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        
        # Mapeo para formateo (Costos y Resultados)
        cost_cols_names = [f'{c} Actual' for c in NOMBRES_COSTOS_INTERNOS] + [f'{c} Antes' for c in NOMBRES_COSTOS_INTERNOS] + ['Result actualizado', 'Resultado anterior']
        cost_cols_indices = [idx + 1 for idx, name in enumerate(header) if name in cost_cols_names]
        
        # AJUSTE CRÍTICO: Formato de Entero para Columnas Específicas
        integer_cols_names = [f'{c} Actual' for c in COLUMNAS_ENTEROS] + [f'{c} Antes' for c in COLUMNAS_ENTEROS]
        integer_cols_indices = [idx + 1 for idx, name in enumerate(header) if name in integer_cols_names]
        integer_format = '#,##0' # Formato para números enteros

        # Mapeo de columnas para el formateo de bordes
        first_col_in_block_names = [f'{output_cost_names.get(c, c)} Actual' for c in NOMBRES_COSTOS_INTERNOS]
        last_col_in_block_names = [f'Impacto {output_cost_names.get(c, c)}' for c in NOMBRES_COSTOS_INTERNOS]
        
        first_col_in_block_indices = [idx + 1 for idx, name in enumerate(header) if name in first_col_in_block_names]
        last_col_in_block_indices = [idx + 1 for idx, name in enumerate(header) if name in last_col_in_block_names]

        # --- Parte 1: Aplicación de Formato de Porcentaje, Moneda y ENTERO ---
        percentage_cols_names = ['% desv', '% parti', '% Variacion Resultado', 'Suma %Parti', 'Suma Impacto']
        for costo in NOMBRES_COSTOS_INTERNOS:
            costo_output = output_cost_names.get(costo, costo)
            percentage_cols_names.append(f'Impacto {costo_output}')

        percentage_indices = [idx + 1 for idx, name in enumerate(header) if name in percentage_cols_names]

        for col_idx in range(1, len(header) + 1):
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    if col_idx in percentage_indices:
                        cell.number_format = '0.00%'
                    elif col_idx in integer_cols_indices:
                        cell.number_format = integer_format # Aplicar formato de entero
                    elif col_idx in cost_cols_indices:
                        # Aplicar formato de moneda a los que no son enteros y a los resultados
                        cell.number_format = currency_format


        # --- Parte 2: Aplicación de Negrita y Color a Encabezados (Fila 1) ---
        for cell in ws[1]:
            col_name = cell.value
            
            if cell.value is not None:
                cell.font = font_black_bold

            if col_name == '% Variacion Resultado':
                cell.fill = fill_variacion_blue
            elif col_name and 'Impacto' in col_name:
                cell.fill = fill_impacto_green
            elif col_name and ('Actual' in col_name or col_name == 'Result actualizado'):
                cell.fill = fill_actual_orange
        
        
        # --- Parte 3: Aplicación de BORDES a TODAS las FILAS ---
        for col_idx in first_col_in_block_indices:
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                current_border = cell.border if cell.border else Border()
                cell.border = Border(left=side_medium, top=current_border.top,  
                                     bottom=current_border.bottom, right=current_border.right)

        for col_idx in last_col_in_block_indices:
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                current_border = cell.border if cell.border else Border()
                cell.border = Border(right=side_medium, top=current_border.top,  
                                     bottom=current_border.bottom, left=current_border.left)
        
        

    except Exception as e:
        st.error(f"❌ Error al aplicar el formato de Excel con openpyxl: {e}")


# --- FUNCIÓN 2: Aplicar Fórmulas Dinámicas al Consolidado (AJUSTADA Y CORREGIDA) ---
def apply_consolidation_formulas(wb, processed_sheet_name, consolidated_sheet_name, df_output_headers, df_consolidado_headers):
    """
    Remplaza los valores estáticos en la hoja consolidada con fórmulas 
    de Excel que referencian a la hoja procesada, asegurando el formato de porcentaje.
    """
    try:
        
        if consolidated_sheet_name not in wb.sheetnames:
            st.warning(f"Advertencia: La hoja '{consolidated_sheet_name}' no se encontró para aplicar fórmulas.")
            return
            
        ws_consolidado = wb[consolidated_sheet_name]
        
        # 1. Mapear la posición de cada columna de salida en la hoja de origen
        header_map = {col_name: idx + 1 for idx, col_name in enumerate(df_output_headers)}
        
        # 2. Definir formato de porcentaje y moneda para el Consolidado
        currency_format = '#,##0'
        percentage_format = '0.00%'
        font_black_bold = Font(color="000000", bold=True)
        fill_impacto_green = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
        fill_variacion_blue = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        
        # 3. Aplicar formato y fórmulas celda por celda
        for col_idx_con, col_name_con in enumerate(df_consolidado_headers):
            
            # Obtener el índice de la columna en la hoja de origen (df_output)
            col_name_source = col_name_con 
            source_col_idx = header_map.get(col_name_source)
            
            if source_col_idx is None:
                st.warning(f"Advertencia: Columna '{col_name_source}' no encontrada en la hoja origen. Se saltará.")
                continue

            source_col_letter = get_column_letter(source_col_idx)
            
            # Aplicar formato de encabezado (negrita y color)
            header_cell = ws_consolidado.cell(row=1, column=col_idx_con + 1)
            header_cell.font = font_black_bold
            if 'Impacto' in col_name_con:
                header_cell.fill = fill_impacto_green
            elif col_name_con == '% Variacion Resultado':
                header_cell.fill = fill_variacion_blue
            
            # Recorrer todas las filas de datos (empezando desde la fila 2)
            for row_idx in range(2, ws_consolidado.max_row + 1):
                cell = ws_consolidado.cell(row=row_idx, column=col_idx_con + 1)
                
                # Crear la fórmula de vinculación
                formula = f"='{processed_sheet_name}'!{source_col_letter}{row_idx}"
                cell.value = formula
                
                # Aplicar formato de número
                # Priorizamos las columnas de porcentaje/impacto para evitar conflictos de formato.
                if col_name_con == '% Variacion Resultado' or 'Impacto' in col_name_con:
                    cell.number_format = percentage_format
                elif 'Result' in col_name_con:
                    cell.number_format = currency_format
                    

    except Exception as e:
        st.error(f"❌ Error al aplicar las fórmulas de Excel con openpyxl: {e}")
        return

# --- FUNCIÓN 3: Escritura y Formateo de la Hoja Procesada con Fórmulas

def write_processed_sheet_with_formulas(wb, sheet_name, df_data, cost_names_internal, output_cost_names, initial_cols):

    try:
        
        # 1. Determinar la posición (índice) de la hoja procesada
        index = 0
        if sheet_name in wb.sheetnames:
            # Obtener el índice actual de la hoja para recrearla en la misma posición
            sheet_names_list = wb.sheetnames
            index = sheet_names_list.index(sheet_name)
            del wb[sheet_name]
        
        # Crear la hoja en su posición original
        ws = wb.create_sheet(sheet_name, index=index)

        # 2. Definir el encabezado final y mapeo de columnas
        header = initial_cols[:]
        for costo in cost_names_internal:
            costo_output = output_cost_names.get(costo, costo)
            header.extend([
                f'{costo_output} Actual',
                f'{costo_output} Antes',
                '% desv',
                '% parti',
                f'Impacto {costo_output}'
            ])
        header.extend(['Result actualizado', 'Resultado anterior', '% Variacion Resultado', 'Suma %Parti', 'Suma Impacto'])
        
        # Escribir el encabezado (Fila 1)
        ws.append(header)
        
        # Crear el mapeo de columnas para las fórmulas
        col_map = {name: get_column_letter(idx + 1) for idx, name in enumerate(header)}
        
        # Mapear las columnas de resultados a sus letras de columna
        result_actual_letter = col_map['Result actualizado']
        result_antes_letter = col_map['Resultado anterior']
        
        # 3. Iterar sobre las filas de datos de pandas e insertar valores/fórmulas
        for row_idx, row in df_data.iterrows():
            excel_row_num = row_idx + 2 # Fila de Excel: 1 (Encabezado) + 1 (Index 0 de Pandas)
            
            current_col_index = 0 # Inicia en la columna 1 (A)
            
            current_parti_cols = []
            current_impacto_cols = []
            
            # --- Bloque Inicial de Columnas (Datos Estáticos) ---
            for col in initial_cols:
                current_col_index += 1
                # Escribir el valor tal cual
                ws.cell(row=excel_row_num, column=current_col_index, value=row[col])

            # --- Bloque de Costos (Valores y Fórmulas) ---
            for costo_interno in cost_names_internal:
                
                # Columna 'Actual' (Valor estático)
                current_col_index += 1
                col_actual = get_column_letter(current_col_index)
                ws.cell(row=excel_row_num, column=current_col_index, value=row[f'{costo_interno} Actual'])
                
                # Columna 'Antes' (Valor estático)
                current_col_index += 1
                col_antes = get_column_letter(current_col_index)
                ws.cell(row=excel_row_num, column=current_col_index, value=row[f'{costo_interno} Antes'])
                
                # Columna '% desv' (Fórmula)
                current_col_index += 1
                cell = ws.cell(row=excel_row_num, column=current_col_index)
                col_desv = get_column_letter(current_col_index)
                # Asegurar el signo de igualdad (=)
                formula_desv = f"=IFERROR(ROUND(({col_actual}{excel_row_num}-{col_antes}{excel_row_num})/{col_antes}{excel_row_num}, 4), 0)"
                cell.value = formula_desv
                
                # Columna '% parti' (Fórmula)
                current_col_index += 1
                cell = ws.cell(row=excel_row_num, column=current_col_index)
                col_parti = get_column_letter(current_col_index)
                # Asegurar el signo de igualdad (=)
                formula_parti = f"=IFERROR(ROUND({col_actual}{excel_row_num}/{result_actual_letter}{excel_row_num}, 4), 0)"
                cell.value = formula_parti
                current_parti_cols.append(f"{col_parti}{excel_row_num}")

                # Columna 'Impacto' (Fórmula)
                current_col_index += 1
                cell = ws.cell(row=excel_row_num, column=current_col_index)
                col_impacto = get_column_letter(current_col_index)
                # Asegurar el signo de igualdad (=)
                formula_impacto = f"=ROUND({col_desv}{excel_row_num}*{col_parti}{excel_row_num}, 4)"
                cell.value = formula_impacto
                current_impacto_cols.append(f"{col_impacto}{excel_row_num}")
                
            # --- Bloque de Resultados (Valores y Fórmulas) ---
            
            # Result actualizado (Valor estático)
            current_col_index += 1
            ws.cell(row=excel_row_num, column=current_col_index, value=row['Result actualizado'])
            
            # Resultado anterior (Valor estático)
            current_col_index += 1
            ws.cell(row=excel_row_num, column=current_col_index, value=row['Resultado anterior'])
            
            # % Variacion Resultado (Fórmula)
            current_col_index += 1
            cell = ws.cell(row=excel_row_num, column=current_col_index)
            # Asegurar el signo de igualdad (=)
            formula_var_res = f"=IFERROR(ROUND(({result_actual_letter}{excel_row_num}-{result_antes_letter}{excel_row_num})/{result_antes_letter}{excel_row_num}, 4), 0)"
            cell.value = formula_var_res
            
            # Suma %Parti (Fórmula)
            current_col_index += 1
            cell = ws.cell(row=excel_row_num, column=current_col_index)
            # Asegurar el signo de igualdad (=)
            formula_suma_parti = f"=ROUND(SUM({'+'.join(current_parti_cols)}), 4)"
            cell.value = formula_suma_parti
            
            # Suma Impacto (Fórmula)
            current_col_index += 1
            cell = ws.cell(row=excel_row_num, column=current_col_index)
            # Asegurar el signo de igualdad (=)
            formula_suma_impacto = f"=ROUND(SUM({'+'.join(current_impacto_cols)}), 4)"
            cell.value = formula_suma_impacto
            
        return header # Retornar el encabezado final para el mapeo del Consolidado
        
    except Exception as e:
        st.error(f"❌ Error al escribir la hoja procesada con fórmulas: {e}")
        return []

# --- FUNCIÓN PRINCIPAL DE PROCESAMIENTO ---
def process_excel_data(uploaded_file):
    
    # 1. Carga del archivo a un objeto BytesIO y a DataFrames de pandas
    try:
        excel_data = uploaded_file.read()
        excel_buffer = io.BytesIO(excel_data)
        
        df_actual = pd.read_excel(excel_buffer, sheet_name=HOJA_ACTUAL, header=0)
        excel_buffer.seek(0)  # MUY IMPORTANTE: resetear el buffer
        df_anterior = pd.read_excel(excel_buffer, sheet_name=HOJA_ANTERIOR, header=0)
        
    except Exception as e:
        st.error(f"❌ ERROR al cargar las hojas de Excel: Asegúrese de que existen las hojas '{HOJA_ACTUAL}' y '{HOJA_ANTERIOR}'. Error: {e}")
        return None, None

    # --- Asignar nombres únicos para manejo interno (Materia vs Material) ---
    column_names_for_df = list(df_actual.columns)
    column_names_for_df = [('Materia_Costo' if col == 'Materia' else col) for col in column_names_for_df]
    
    df_actual.columns = column_names_for_df
    df_anterior.columns = column_names_for_df
    
    # --- 3. Preparación de columnas para la combinación ---
    df_actual = df_actual.rename(columns={COLUMNA_RESULTADO: 'Result actualizado'})
    df_anterior = df_anterior.rename(columns={COLUMNA_RESULTADO: 'Resultado anterior'})
    
    rename_actual = {col: f'{col} Actual' for col in NOMBRES_COSTOS_INTERNOS}
    rename_anterior = {col: f'{col} Antes' for col in NOMBRES_COSTOS_INTERNOS}
    
    df_actual_renamed = df_actual.rename(columns=rename_actual).copy()
    df_anterior_renamed = df_anterior.rename(columns=rename_anterior).copy()
    
    cols_to_keep_anterior = [CLAVE_MERGE, 'Resultado anterior'] + list(rename_anterior.values())
    df_anterior_slim = df_anterior_renamed[cols_to_keep_anterior]
    
    # --- 4. Combinación (Merge) ---
    df_procesado = pd.merge(df_actual_renamed, df_anterior_slim, on=CLAVE_MERGE, how='left')
    
    # --- 5. Preparación de datos y columnas a mantener (Aplicación de redondeo) ---
    cols_to_keep = ['Versi', 'Ce.', CLAVE_MERGE, 'Texto breve material', 'Pr', 'UMB', 'Válido de', 'Tam.lot', 'Costo d', 'Result actualizado', 'Resultado anterior']
    
    for costo in NOMBRES_COSTOS_INTERNOS:
        cols_to_keep.append(f'{costo} Actual')
        cols_to_keep.append(f'{costo} Antes')
    
    df_input_for_excel = df_procesado[cols_to_keep].copy()
    
    # Convertir columnas numéricas que se usarán en fórmulas (asegurar el tipo)
    cols_numeric = [f'{c} Actual' for c in NOMBRES_COSTOS_INTERNOS] + [f'{c} Antes' for c in NOMBRES_COSTOS_INTERNOS] + ['Result actualizado', 'Resultado anterior']
    
    # APLICACIÓN DE REDONDEO SOLICITADO
    cols_integer_actual = [f'{c} Actual' for c in COLUMNAS_ENTEROS]
    cols_integer_antes = [f'{c} Antes' for c in COLUMNAS_ENTEROS]
    
    for col in cols_numeric:
        col_series = pd.to_numeric(df_input_for_excel[col], errors='coerce')
        if col in cols_integer_actual or col in cols_integer_antes:
            # Redondeo a 0 decimales (Entero)
            df_input_for_excel[col] = col_series.round(0).fillna(0).astype(int) 
        else:
            # Redondeo a 2 decimales para los demás (manteniendo la lógica original)
            df_input_for_excel[col] = col_series.round(2)
            
    # -------------------------------------------------------------------------------------
    # --- 6. Guardar y Formatear las hojas en un objeto de memoria ---
    # -------------------------------------------------------------------------------------
    
    output_file = io.BytesIO()
    
    try:
        # Cargamos el archivo original en el Workbook de openpyxl
        wb = load_workbook(io.BytesIO(excel_data))
        
        initial_cols = ['Versi', 'Ce.', CLAVE_MERGE, 'Texto breve material', 'Pr', 'UMB', 'Válido de', 'Tam.lot', 'Costo d']
        
        # 6.1 Escritura de la Hoja PROCESADA con FÓRMULAS
        df_output_headers = write_processed_sheet_with_formulas(wb, HOJA_PROCESADA, df_input_for_excel, NOMBRES_COSTOS_INTERNOS, output_cost_names, initial_cols)
        
        if not df_output_headers:
            st.error("\nEl script se detuvo debido a un error al escribir la hoja procesada.")
            return None, None
            
        
        
        # 6.2 Aplicar formato a la hoja de PROCESADO
        apply_excel_formatting(wb, HOJA_PROCESADA)

        # -------------------------------------------------------------------------------------
        # --- 6.5. PREPARACIÓN Y ESCRITURA DEL CONSOLIDADO 
        # -------------------------------------------------------------------------------------
        
        cols_consolidado = [CLAVE_MERGE, 'Texto breve material', 'Result actualizado', 'Resultado anterior', '% Variacion Resultado']
        
        for costo_interno in NOMBRES_COSTOS_INTERNOS:
            costo_output = output_cost_names.get(costo_interno, costo_interno)
            cols_consolidado.append(f'Impacto {costo_output}')
        
        df_consolidado_headers = cols_consolidado 
        
        # Guardar la hoja Consolidado con el tamaño correcto usando openpyxl
        num_rows = len(df_input_for_excel)
        
        index_consolidado = 0
        if HOJA_CONSOLIDADO in wb.sheetnames:
            index_consolidado = wb.sheetnames.index(HOJA_CONSOLIDADO)
            del wb[HOJA_CONSOLIDADO]
            
        ws_consolidado = wb.create_sheet(HOJA_CONSOLIDADO, index=index_consolidado)
        
        # Escribir el encabezado
        ws_consolidado.append(df_consolidado_headers)
        # Llenar con celdas dummy (el contenido será reemplazado por fórmulas)
        for _ in range(num_rows):
            ws_consolidado.append([0] * len(df_consolidado_headers))

        
        # 6.3 Aplicar FÓRMULAS VINCULANTES y formato a la hoja de CONSOLIDADO
        apply_consolidation_formulas(wb, HOJA_PROCESADA, HOJA_CONSOLIDADO, df_output_headers, df_consolidado_headers)

        # Guardar el libro de trabajo en el buffer de memoria
        wb.save(output_file)
        
        st.success("\n¡El script ha terminado exitosamente! Las hojas ahora contienen fórmulas Excel y el orden original se ha mantenido.")
        
        return output_file, uploaded_file.name.replace(".xlsx", "_PROCESADO.xlsx")

    except Exception as e:
        st.error(f"❌ Ocurrió un error inesperado durante el procesamiento final: {e}")
        return None, None

# --- Interfaz de Streamlit ---
st.set_page_config(
    page_title="Procesador de Costos Excel",
    layout="centered",
    initial_sidebar_state="auto"
)

st.title("📊 Automatización: Cálculo de la variación de costos")
st.markdown("---")

st.info("""
    Este programa lee las hojas **'Costos Mqlla H5 ACTUAL'** y **'Costos Mqlla H5 ANTERIOR'** del archivo Excel cargado.

    Realiza el cálculo de desvío, participación e impacto entre los costos, y genera dos hojas de salida:
    1.  **'Costos_procesado'**: Con todos los detalles y valores estáticos para los campos establecidos
    2.  **'Consolidado_Impactos'**: Resumen con los resultados finales.

    ⚠️ **Importante**: Asegúrese de que las hojas de origen y destino existan en el archivo original.
""")

uploaded_file = st.file_uploader(
    "📤 Seleccione el archivo Excel de Costos:",
    type=["xlsx"],
    accept_multiple_files=False
)

if uploaded_file is not None:
    st.success(f"Archivo cargado: **{uploaded_file.name}**")
    
    if st.button("🚀 Iniciar Procesamiento y Formateo"):
        with st.spinner("Procesando datos, generando fórmulas y aplicando formato..."):
            output_buffer, output_filename = process_excel_data(uploaded_file)
        
        st.markdown("---")

        if output_buffer:
            st.download_button(
                label="📥 Descargar Archivo Excel Procesado",
                data=output_buffer.getvalue(),
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        else:

            st.error("El procesamiento falló. Revise los mensajes de error anteriores.")
